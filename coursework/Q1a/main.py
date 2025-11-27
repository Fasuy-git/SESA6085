from pathlib import Path
from openpyxl import load_workbook
import numpy as np
import scipy
import time


def f_1(t, mu, sigma):
    """PDF of the log normal distribution"""
    front_coefficient = 1 / (t * sigma * np.sqrt(2 * np.pi))
    exponent = np.exp(-0.5 * ((np.log(t) - mu) / (sigma)) ** 2)
    return front_coefficient * exponent


def f_2(t, mu, sigma):
    """PDF of the normal distribution"""
    front_coefficient = 1 / (sigma * np.sqrt(2 * np.pi))
    exponent = np.exp(-0.5 * ((t - mu) / (sigma)) ** 2)
    return front_coefficient * exponent


def F_1(t, mu, sigma):
    """CDF of the log normal distribution"""
    z = (np.log(t) - mu) / (sigma * np.sqrt(2))
    error_function = scipy.special.erf(z)
    return 0.5 * (1 + error_function)


def F_2(t, mu, sigma):
    """CDF of the normal distribution"""
    z = (t - mu) / (sigma * np.sqrt(2))
    error_function = scipy.special.erf(z)
    return 0.5 * (1 + error_function)


def log_likelihood(mu_1, mu_2, sigma_1, sigma_2, extracted_excel_data):
    """Calculated the log likelihood of the inputted censored data"""
    T_R, C_R, T_L, C_L = extracted_excel_data[1:5]
    failure_time_data = np.array(extracted_excel_data[5:])

    # Calculate uncensored data likelihood
    log_normal_pdf = f_1(failure_time_data, mu_1, sigma_1)
    normal_pdf = f_2(failure_time_data, mu_2, sigma_2)

    log_normal_reliability = 1 - F_1(failure_time_data, mu_1, sigma_1)
    normal_reliability = 1 - F_2(failure_time_data, mu_2, sigma_2)

    uncensored_data_total = np.sum(
        np.log(log_normal_pdf * normal_reliability + normal_pdf * log_normal_reliability)
    )

    # Calculate left censored data likelihood
    RT1_L = 1 - F_1(T_L, mu_1, sigma_1)
    RT2_L = 1 - F_2(T_L, mu_2, sigma_2)
    left_censored_total = C_L * np.log(1 - (RT1_L * RT2_L))

    # Calculate right censored data likelihood
    RT1_R = 1 - F_1(T_R, mu_1, sigma_1)
    RT2_R = 1 - F_2(T_R, mu_2, sigma_2)
    right_censored_total = C_R * np.log(RT1_R * RT2_R)

    # Compute and return log likelihood
    return uncensored_data_total + left_censored_total + right_censored_total


if __name__ == "__main__":
    # start time
    start = time.time()

    # Define Constants
    STUDENT_ID = 32750552
    N = 100000  # Number of trails

    # Extract failure data from excel file
    # Build excel file loc
    excel_file_dir = Path(__file__).resolve().parent.parent
    excel_file_loc = Path.joinpath(excel_file_dir, "Q1_Data.xlsx")

    # Open workbook and extract data
    work_book = load_workbook(excel_file_loc)
    sheet = work_book.active

    # Find relevant row in excel spreadsheet
    for row in sheet.iter_rows(values_only=True):
        if row[0] == STUDENT_ID:
            extracted_excel_data = row
    work_book.close()

    # Define limits from coursework instructions
    lower_limits = np.array([1.0, 0.5, 20.0, 0.5])
    upper_limits = np.array([5.0, 5.0, 60.0, 10.0])

    # Create N points for each parameter within bounds
    mu1_vals = np.linspace(lower_limits[0], upper_limits[0], N)
    s1_vals = np.linspace(lower_limits[1], upper_limits[1], N)
    mu2_vals = np.linspace(lower_limits[2], upper_limits[2], N)
    s2_vals = np.linspace(lower_limits[3], upper_limits[3], N)

    best_ll = -np.inf
    best_params = None

    start = time.time()

    # Brute-force all 4-D combinations using vectorized functions
    for mu1, s1, mu2, s2 in zip(mu1_vals, s1_vals, mu2_vals, s2_vals):  # MC or QMC samples
        ll = log_likelihood(mu1, mu2, s1, s2, extracted_excel_data)
        if ll > best_ll:
            best_ll = ll
            best_params = (mu1, s1, mu2, s2)

    end = time.time()

    print("Best log-likelihood:", best_ll)
    print("Best params (mu_1, sigma_1, mu_2, sigma):", best_params)
    print(f"Elapsed time {end - start:.3f} seconds")
